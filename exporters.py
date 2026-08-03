"""
导出模块：商品数据导出为 CSV、XLSX、标准 JSON (数组)
"""
import csv
import json
import logging
import os
from threading import Lock

from models import ProductItem
from utils import write_atomic_json

logger = logging.getLogger("emag_crawler.exporters")


class Exporters:
    """商品数据导出器, 线程安全"""

    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        self.csv_path = os.path.join(output_dir, "products.csv")
        self.xlsx_path = os.path.join(output_dir, "products.xlsx")
        self.json_path = os.path.join(output_dir, "products.json")

        # 所有商品数据 (list of dict)
        self._products: list[dict] = []
        self._lock = Lock()

    def add_product(self, product: ProductItem) -> None:
        """添加一个商品 (线程安全)"""
        d = _product_to_json_dict(product)
        with self._lock:
            self._products.append(d)

    def add_products(self, products: list[ProductItem]) -> None:
        """批量添加商品 (线程安全)"""
        dicts = [_product_to_json_dict(p) for p in products]
        with self._lock:
            self._products.extend(dicts)

    def get_products_sorted(self) -> list[dict]:
        """按类目、页码、页面位置排序后返回"""
        with self._lock:
            sorted_list = sorted(
                self._products,
                key=lambda x: (
                    x.get("category_name", ""),
                    x.get("page_number", 0),
                    x.get("position_in_page", 0),
                ),
            )
            return sorted_list

    def get_product_count(self) -> int:
        """已写入商品数"""
        with self._lock:
            return len(self._products)

    def get_csv_buffer(self) -> list[dict]:
        """获取用于 CSV/XLSX 导出的数据 (extra 转为 JSON 字符串)"""
        sorted_prods = self.get_products_sorted()
        result = []
        for d in sorted_prods:
            row = dict(d)
            if isinstance(row.get("extra"), dict):
                row["extra"] = json.dumps(row["extra"], ensure_ascii=False) if row["extra"] else ""
            result.append(row)
        return result

    def write_json(self) -> None:
        """写入标准 products.json (JSON 数组, 原子写入)"""
        sorted_prods = self.get_products_sorted()
        # extra 在 JSON 中保持为 object
        write_atomic_json(self.json_path, sorted_prods)
        logger.info(f"JSON 已写入: {self.json_path} ({len(sorted_prods)} 条)")

    def _write_csv(self) -> None:
        """写入 CSV (UTF-8 BOM)"""
        buffer = self.get_csv_buffer()
        if not buffer:
            # 至少创建表头
            columns = ProductItem.csv_columns()
            field_names = ProductItem.field_names()
            with open(self.csv_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=field_names, extrasaction="ignore")
                writer.writerow(dict(zip(field_names, columns)))
            logger.info(f"CSV 已写入(仅表头): {self.csv_path}")
            return

        columns = ProductItem.csv_columns()
        field_names = ProductItem.field_names()
        with open(self.csv_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=field_names, extrasaction="ignore")
            writer.writerow(dict(zip(field_names, columns)))
            for row in buffer:
                writer.writerow(row)
        logger.info(f"CSV 已写入: {self.csv_path} ({len(buffer)} 行)")

    def _write_xlsx(self) -> None:
        """写入 XLSX (带格式化)"""
        buffer = self.get_csv_buffer()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl 未安装，跳过 XLSX")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "商品数据"

        columns = ProductItem.excel_columns()
        header_fill = PatternFill(start_color="005EB8", end_color="005EB8", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")

        for col_idx, (col_name, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        if buffer:
            price_fields = {"price_current", "price_old", "price_promo"}
            for row_idx, item in enumerate(buffer, 2):
                for col_idx, (_, field_name) in enumerate(columns, 1):
                    value = item.get(field_name, "")
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if field_name in price_fields and value is not None and value != "":
                        try:
                            cell.value = float(value)
                            cell.number_format = '#,##0.00'
                        except (ValueError, TypeError):
                            cell.value = str(value) if value is not None else ""
                    elif field_name in ("discount_percent", "rating", "review_count",
                                        "page_number", "position_in_page", "http_status"):
                        try:
                            cell.value = int(value) if value is not None and value != "" else value
                        except (ValueError, TypeError):
                            cell.value = str(value) if value is not None else ""
                    else:
                        cell.value = str(value) if value is not None else ""

        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(len(buffer) + 1, 2)}"
        ws.freeze_panes = "A2"

        col_widths = {1: 16, 10: 45, 11: 65, 12: 12, 13: 18, 14: 12, 15: 18, 24: 30, 28: 70, 29: 50, 30: 22, 33: 40}
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        wb.save(self.xlsx_path)
        logger.info(f"XLSX 已写入: {self.xlsx_path} ({len(buffer)} 行)")

    def finalize(self) -> None:
        """完成导出"""
        self.write_json()
        self._write_csv()
        self._write_xlsx()


def _product_to_json_dict(product: ProductItem) -> dict:
    """将 ProductItem 转为适合 JSON 输出的 dict (extra 保持为 object)"""
    from dataclasses import asdict
    d = asdict(product)
    # JSON 中 extra 保持为原始 dict, 不被二次编码
    if not d.get("extra"):
        d.pop("extra", None)
    return d
